#!/usr/bin/env python

import pandas as pd
from openpyxl import load_workbook
import re
import itertools
import string
import uuid
import json
from pathlib import Path

source = './dummy.xlsx'
metadata = ['Submission Date']
metadata_prefix = ["#"]
ignore = ["datapoint"]
Path("./data").mkdir(parents=True, exist_ok=True)


def excel_cols():
    n = 1
    while True:
        azstring = itertools.product(string.ascii_uppercase, repeat=n)
        azstring = (''.join(group) for group in azstring)
        yield from azstring
        n += 1


def find_unit(c):
    u = re.search(r'\((.*?)\)', c)
    if u:
        return u.group(1)
    return False


def trim_uuid(x):
    return "-".join(str(x).split("-")[1:-1])


all_sheets = load_workbook(source, read_only=True).sheetnames
sheets = list(filter(lambda x: '|' in x, all_sheets))
registration = list(filter(lambda x: 'registration' in x, sheets))[0]
regdf = pd.read_excel(source, registration)
uuid_list = [uuid.uuid4() for _ in range(len(regdf.index))]

settings = list(filter(lambda x: 'settings' in x, all_sheets))[0]
settings = pd.read_excel(source, settings)
settings['question'] = settings['question'].apply(
    lambda x: x.replace('\n', "").strip())
settings['table'] = settings['table'].apply(lambda x: x[0:31])

guides = list(filter(lambda x: 'guides' == x, all_sheets))[0]
guides = pd.read_excel(source, guides)
guides['question'] = guides['question'].apply(
    lambda x: x.split('(')[0].strip())
guides['note'] = guides['note'].apply(lambda x: None if x != x else x)
guides = guides.to_dict('records')
with open('./data/guides.json', 'w') as f:
    json.dump(guides, f)

static_list = list(filter(lambda x: 'static' in x, all_sheets))
statics = []
for static in static_list:
    stdict = pd.read_excel(source, static)
    table_name = stdict['table'][0].title()
    stdict = stdict.drop(columns='table').to_dict('records')
    for st in stdict:
        for s in st:
            if type(st[s]) == str:
                st.update({s: st[s].replace("\n", "").strip()})
            if st[s] != st[s]:
                st.update({s: None})
    statics.append({'table': table_name, 'data': stdict})
with open('./data/statics.json', 'w') as f:
    json.dump(statics, f)

configs = []
for tab in sheets:
    definitions = []
    col_rename = {}
    sheet = tab.split('|')
    sheet = {
        'type': sheet[0],
        'name': sheet[1], 'file': sheet[1].lower().replace(' ', '_') }
    df = pd.read_excel(source, tab)
    df['uuid'] = df['datapoint'].apply(lambda x: uuid_list[int(x) - 1])
    df['uuid'] = df['uuid'].apply(lambda x: trim_uuid(x))
    df = df.drop(columns=ignore, axis=1)
    strings = list(itertools.islice(excel_cols(), df.shape[1]))
    for idx, col in enumerate(list(df.columns)):
        col_name = col.replace('\n', "").strip()
        name = col_name.replace('#', '')
        dtype = str(df[col].dtypes.type)
        dtype = re.search(r'\'(.*?)\'', dtype).group(0)
        dtype = dtype.replace("'", "").replace('numpy.', '')
        unit = find_unit(name)
        name = name.split(' (')[0]
        name = name.strip().replace('_', ' ').title()
        meta = False
        if col in metadata:
            meta = True
        st = settings[(settings['table'] == tab)
                      & (settings['question'] == col)]
        options = False
        if st.shape[0]:
            options = st[['option', 'color']].to_dict('records')
        definitions.append({
            'name': name,
            'alias': strings[idx],
            'original': col,
            'unit': unit,
            'meta': meta,
            'dtype': dtype,
            'options': options
        })
        col_rename.update({col: strings[idx]})
    df = df.rename(columns=col_rename)
    df = df.apply(lambda x: x.fillna(0)
                  if x.dtype.kind in 'biufc' else x.fillna("-"))
    result = './data/{}.csv'.format(sheet['type'])
    if tab != registration:
        Path("./data/{}".format(sheet['type'])).mkdir(parents=True,
                                                      exist_ok=True)
        result = "./data/{}/{}.csv".format(sheet['type'], sheet['file'])
    df.to_csv(result, index=False)
    sheet.update({'definition': definitions})
    configs.append(sheet)
with open('./data/configs.json', 'w') as f:
    json.dump(configs, f)
